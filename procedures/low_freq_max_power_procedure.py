import time
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from base_test_procedure import BaseTestProcedure, format_frequency


class LowFreqMaxPowerProcedure(BaseTestProcedure):
    """低频段最大功率测试流程类"""

    def __init__(self, instrument_manager):
        """初始化测试流程

        Args:
            instrument_manager: 仪器管理器对象
        """
        super().__init__(instrument_manager)
        self.power_sweep_data = []
        self.csv_sweep_streamer = None

    def add_test_result(self, frequency, max_power, max_measured_power, 
                       attenuation, saturation_point, steps, notes):
        """添加测试结果"""
        self.test_results.append({
            'frequency': frequency,
            'max_power': max_power,
            'max_measured_power': max_measured_power,
            'attenuation': attenuation,
            'saturation_point': saturation_point,
            'steps': steps,
            'notes': notes,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    def add_power_sweep_point(self, frequency, set_power, measured_power, 
                             actual_power, step_index, status):
        """添加功率扫描数据点"""
        self.power_sweep_data.append({
            'frequency': frequency,
            'set_power': set_power,
            'measured_power': measured_power,
            'actual_power': actual_power,
            'step_index': step_index,
            'status': status,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        if self.csv_sweep_streamer:
            self.csv_sweep_streamer.append(self.power_sweep_data[-1])

    def run_test(self, signal_generator, spectrum_analyzer, test_config, keep_output=False):
        """运行低频段最大功率测试"""
        frequency = test_config['frequency']
        start_power = test_config['start_power']
        power_step = test_config['power_step']
        max_set_power = test_config['max_set_power']
        max_measured_power = test_config['max_measured_power']
        power_tolerance = test_config['power_tolerance']
        max_power_drop = test_config['max_power_drop']
        attenuator_value = test_config['attenuator_value']
        use_attenuator = test_config['use_attenuator']
        settling_time = test_config['settling_time']
        sa_settling_time = test_config['sa_settling_time']
        measurement_times = test_config['measurement_times']
        sa_config = test_config['spectrum_analyzer_config']

        print(f"\n{'=' * 60}")
        print(f"开始低频段最大功率测试: {test_config['test_name']}")
        print(f"频率: {format_frequency(frequency)}")
        print(f"起始功率: {start_power} dBm, 步进: {power_step} dB")
        print(f"最大设定功率限制: {max_set_power} dBm")
        print(f"最大测量功率限制: {max_measured_power} dBm")
        if use_attenuator:
            print(f"衰减器值: {attenuator_value} dB")
        print(f"频谱仪输入耦合: {sa_config['input_coupling']}")
        print(f"{'=' * 60}")

        signal_generator.set_frequency(frequency)
        
        spectrum_analyzer.set_center_frequency(frequency)
        spectrum_analyzer.set_span(sa_config['span'])
        spectrum_analyzer.set_rbw(sa_config['rbw'])
        spectrum_analyzer.set_vbw(sa_config['vbw'])
        spectrum_analyzer.set_reference_level(sa_config['reference_level'])
        spectrum_analyzer.set_attenuation(sa_config['attenuation'])
        spectrum_analyzer.set_input_coupling(sa_config['input_coupling'])
        print(f"设置频谱仪输入耦合为: {sa_config['input_coupling']}")
        
        current_power = start_power
        signal_generator.set_power(current_power)
        signal_generator.enable_output(True)
        print(f"启用信号源输出，起始功率: {current_power} dBm")
        
        print(f"等待信号源稳定 {settling_time}秒...")
        time.sleep(settling_time)
        print(f"等待频谱仪稳定 {sa_settling_time}秒...")
        time.sleep(sa_settling_time)
        
        max_achieved_power = None
        max_achieved_measured = None
        prev_measured_power = None
        saturation_detected = False
        overload_detected = False
        step_count = 0
        stop_reason = "正常完成"
        
        while True:
            step_count += 1
            print(f"\n--- 功率步进 {step_count}: 设定功率 = {current_power:.1f} dBm ---")
            
            signal_generator.set_power(current_power)
            time.sleep(settling_time)
            
            measured_power = None
            measurements = []
            for i in range(measurement_times):
                power = spectrum_analyzer.measure_power()
                if power is not None:
                    measurements.append(power)
                time.sleep(0.1)
            
            if measurements:
                measured_power = sum(measurements) / len(measurements)
                print(f"测量功率 (平均{len(measurements)}次): {measured_power:.2f} dBm")
            else:
                print("警告: 频谱仪测量失败，跳过此点")
                self.add_power_sweep_point(frequency, current_power, None, None, step_count, '测量失败')
                measured_power = prev_measured_power if prev_measured_power is not None else -float('inf')
            
            if use_attenuator:
                actual_power = measured_power + attenuator_value
            else:
                actual_power = measured_power
            
            print(f"实际功率: {actual_power:.2f} dBm")
            
            stop_scan = False
            
            if current_power >= max_set_power:
                stop_reason = f"达到设定功率限制 ({max_set_power} dBm)"
                print(f"停止条件: {stop_reason}")
                stop_scan = True
            
            if measured_power >= max_measured_power:
                stop_reason = f"达到测量功率限制 ({max_measured_power} dBm)"
                print(f"停止条件: {stop_reason}")
                stop_scan = True
            
            if prev_measured_power is not None:
                power_increase = measured_power - prev_measured_power
                if power_increase < power_tolerance:
                    saturation_detected = True
                    stop_reason = f"检测到饱和 (功率增加仅{power_increase:.2f} dB < 容差 {power_tolerance} dB)"
                    print(f"停止条件: {stop_reason}")
                    stop_scan = True
            
            if prev_measured_power is not None:
                power_drop = prev_measured_power - measured_power
                if power_drop > max_power_drop:
                    overload_detected = True
                    stop_reason = f"检测到过载 (功率下降 {power_drop:.2f} dB > 最大允许 {max_power_drop} dB)"
                    print(f"停止条件: {stop_reason}")
                    stop_scan = True
            
            status = '正常'
            if saturation_detected:
                status = '饱和'
            elif overload_detected:
                status = '过载'
            elif stop_scan:
                status = '超限'
            
            self.add_power_sweep_point(frequency, current_power, measured_power, actual_power, step_count, status)
            
            if max_achieved_power is None or actual_power > max_achieved_power:
                max_achieved_power = actual_power
                max_achieved_measured = measured_power
            
            prev_measured_power = measured_power
            
            if stop_scan:
                break
            
            current_power += power_step
        
        notes = stop_reason
        if saturation_detected:
            notes += " (饱和点)"
        if overload_detected:
            notes += " (过载点)"
        
        self.add_test_result(
            frequency=frequency,
            max_power=max_achieved_power,
            max_measured_power=max_achieved_measured,
            attenuation=attenuator_value if use_attenuator else 0.0,
            saturation_point=saturation_detected,
            steps=step_count,
            notes=notes
        )
        
        print(f"\n{'=' * 60}")
        print(f"测试完成: {test_config['test_name']}")
        print(f"最大实际功率: {max_achieved_power:.2f} dBm")
        print(f"最大测量功率: {max_achieved_measured:.2f} dBm")
        print(f"功率步进数: {step_count}")
        print(f"停止原因: {stop_reason}")
        print(f"{'=' * 60}")
        
        if not keep_output:
            signal_generator.enable_output(False)
            print("信号源输出已禁用")
            time.sleep(test_config['post_close_wait'])
        else:
            print("保持信号源输出状态")

    def start_csv_stream(self, csv_path):
        """开启 CSV 流式写入"""
        from utils.csv_streamer import CsvStreamer
        self.csv_sweep_streamer = CsvStreamer(
            csv_path.replace(".csv", "_sweep.csv"), [
                "frequency", "set_power", "measured_power", "actual_power",
                "step_index", "status", "timestamp",
            ]
        )

    def finish_xlsx(self, xlsx_path):
        """关闭 CSV 流并转为 XLSX"""
        if not self.csv_sweep_streamer:
            return self.save_results(xlsx_path)
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment
        df_sweep = pd.read_csv(self.csv_sweep_streamer.filepath, encoding="utf-8-sig")
        self.csv_sweep_streamer.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "最大功率摘要"
        for col, h in enumerate(["频率 (Hz)", "最大实际功率 (dBm)", "最大测量功率 (dBm)", "衰减值 (dB)", "是否饱和", "步进数", "备注", "时间戳"], 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for r, row in pd.DataFrame(self.test_results).iterrows():
            for c, val in enumerate(row, 1):
                ws.cell(row=r+4, column=c, value=val)
        ws2 = wb.create_sheet(title="详细功率扫描数据")
        for col, h in enumerate(["频率 (Hz)", "设定功率 (dBm)", "测量功率 (dBm)", "实际功率 (dBm)", "步进索引", "状态", "时间戳"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for r, row in df_sweep.iterrows():
            for c, val in enumerate(row, 1):
                ws2.cell(row=r+2, column=c, value=val)
        wb.save(xlsx_path)
        print(f"Excel 已保存: {xlsx_path}")

    def save_results(self, filename):
        """保存测试结果"""
        if not self.test_results:
            print("没有测试结果可保存")
            return

        try:
            if filename.endswith('.xlsx'):
                wb = openpyxl.Workbook()
                ws_summary = wb.active
                ws_summary.title = "最大功率摘要"
                
                headers = ['频率 (Hz)', '频率显示', '最大实际功率 (dBm)', '最大测量功率 (dBm)', 
                          '衰减值 (dB)', '是否饱和', '功率步进数', '备注', '时间戳']
                for col, header in enumerate(headers, 1):
                    cell = ws_summary.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                for row, result in enumerate(self.test_results, 2):
                    ws_summary.cell(row=row, column=1).value = result['frequency']
                    ws_summary.cell(row=row, column=2).value = format_frequency(result['frequency'])
                    ws_summary.cell(row=row, column=3).value = result['max_power']
                    ws_summary.cell(row=row, column=4).value = result['max_measured_power']
                    ws_summary.cell(row=row, column=5).value = result['attenuation']
                    ws_summary.cell(row=row, column=6).value = '是' if result['saturation_point'] else '否'
                    ws_summary.cell(row=row, column=7).value = result['steps']
                    ws_summary.cell(row=row, column=8).value = result['notes']
                    ws_summary.cell(row=row, column=9).value = result['timestamp']
                
                for col in range(1, len(headers) + 1):
                    ws_summary.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                ws_detail = wb.create_sheet(title="详细功率扫描数据")
                detail_headers = ['频率 (Hz)', '设定功率 (dBm)', '测量功率 (dBm)', 
                                '实际功率 (dBm)', '步进索引', '状态', '时间戳']
                for col, header in enumerate(detail_headers, 1):
                    cell = ws_detail.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                for row, data in enumerate(self.power_sweep_data, 2):
                    ws_detail.cell(row=row, column=1).value = data['frequency']
                    ws_detail.cell(row=row, column=2).value = data['set_power']
                    ws_detail.cell(row=row, column=3).value = data['measured_power']
                    ws_detail.cell(row=row, column=4).value = data['actual_power']
                    ws_detail.cell(row=row, column=5).value = data['step_index']
                    ws_detail.cell(row=row, column=6).value = data['status']
                    ws_detail.cell(row=row, column=7).value = data['timestamp']
                
                for col in range(1, len(detail_headers) + 1):
                    ws_detail.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                wb.save(filename)
                print(f"测试结果已保存到Excel文件: {filename}")
            else:
                base_name = filename.replace('.csv', '')
                summary_file = f"{base_name}_summary.csv"
                detail_file = f"{base_name}_detail.csv"
                
                with open(summary_file, 'w', newline='') as csvfile:
                    fieldnames = ['frequency', 'max_power', 'max_measured_power', 
                                 'attenuation', 'saturation_point', 'steps', 'notes', 'timestamp']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for result in self.test_results:
                        writer.writerow(result)
                
                with open(detail_file, 'w', newline='') as csvfile:
                    fieldnames = ['frequency', 'set_power', 'measured_power', 
                                 'actual_power', 'step_index', 'status', 'timestamp']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for data in self.power_sweep_data:
                        writer.writerow(data)
                
                print(f"测试结果已保存到CSV文件:")
                print(f"  摘要: {summary_file}")
                print(f"  详细数据: {detail_file}")
        except Exception as e:
            print(f"保存测试结果失败: {e}")