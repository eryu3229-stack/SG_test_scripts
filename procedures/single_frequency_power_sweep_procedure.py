#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""单频点功率扫描流程（功率计版本）"""

import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font


def format_frequency(frequency):
    """根据频率大小自动选择合适的单位显示。"""
    if frequency < 1e3:
        return f"{frequency:.0f}Hz"
    if frequency < 1e6:
        return f"{frequency / 1e3:.2f}kHz"
    if frequency < 1e9:
        return f"{frequency / 1e6:.2f}MHz"
    return f"{frequency / 1e9:.2f}GHz"


class SingleFrequencyPowerSweepProcedure:
    """在固定频点上扫描设定功率，并用功率计逐点测量实际功率。"""

    def __init__(self, instrument_manager):
        self.instrument_manager = instrument_manager
        self.test_results = []
        self.power_sweep_data = []

    def add_power_sweep_point(self, frequency, set_power, measured_power,
                              compensated_power, status):
        """添加单个功率扫描点。"""
        self.power_sweep_data.append({
            'frequency': frequency,
            'set_power': set_power,
            'measured_power': measured_power,
            'compensated_power': compensated_power,
            'status': status,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })

    def add_test_result(self, frequency, total_points, measured_points,
                        max_measured_power, min_measured_power,
                        attenuator_enabled, attenuation, stop_reason):
        """添加测试摘要结果。"""
        self.test_results.append({
            'frequency': frequency,
            'total_points': total_points,
            'measured_points': measured_points,
            'max_measured_power': max_measured_power,
            'min_measured_power': min_measured_power,
            'attenuator_enabled': attenuator_enabled,
            'attenuation': attenuation,
            'stop_reason': stop_reason,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })

    def _measure_with_retry(self, power_meter, measurement_times):
        """测量功率，失败时最多重试 3 次。"""
        for attempt in range(1, 4):
            measured_power = power_meter.measure_power(times=measurement_times)
            if measured_power is not None:
                return measured_power
            if attempt < 3:
                print(f"测量失败，0.5 秒后重试（第 {attempt + 1}/3 次）...")
                time.sleep(0.5)
        return None

    def run_test(self, signal_generator, power_meter, test_config, keep_output=False):
        """运行单频点功率扫描。"""
        frequency = test_config['frequency']
        max_set_power = test_config['max_set_power']
        max_measured_power = test_config['max_measured_power']
        settling_time = test_config['settling_time']
        pm_settling_time = test_config['pm_settling_time']
        post_close_wait = test_config['post_close_wait']
        measurement_times = test_config['measurement_times']
        attenuator_enabled = test_config['attenuator_enabled']
        attenuator_value = test_config['attenuator_value']

        power_points = [
            power for power in test_config['power_points']
            if power <= max_set_power
        ]
        if not power_points:
            print("配置错误：没有落在 max_set_power 内的功率扫描点，请检查配置。")
            return
        if len(power_points) < len(test_config['power_points']):
            print(f"提示：{len(test_config['power_points']) - len(power_points)} "
                  f"个功率点超过 max_set_power，已跳过。")

        print(f"\n{'=' * 60}")
        print(f"开始单频点功率扫描: {test_config['test_name']}")
        print(f"频率: {format_frequency(frequency)}")
        print(f"功率范围: {power_points[0]} ~ {power_points[-1]} dBm, "
              f"共 {len(power_points)} 个点")
        if attenuator_enabled:
            print(f"衰减器补偿: {attenuator_value} dB")
        print(f"{'=' * 60}")

        signal_generator.set_frequency(frequency)
        power_meter.set_frequency(frequency)
        time.sleep(pm_settling_time)

        if attenuator_enabled and hasattr(power_meter, 'set_input_attenuation'):
            power_meter.set_input_attenuation(attenuator_value)
            print(f"设置功率计输入衰减: {attenuator_value} dB")

        first_power = power_points[0]
        signal_generator.set_power(first_power)
        signal_generator.enable_output(True)
        print(f"启用信号源输出，起始功率: {first_power} dBm")
        time.sleep(settling_time)

        consecutive_failures = 0
        measured_count = 0
        max_measured = None
        min_measured = None
        stop_reason = "扫描完成"

        for index, set_power in enumerate(power_points, 1):
            print(f"\n--- 功率点 {index}/{len(power_points)}: "
                  f"设定功率 = {set_power} dBm ---")
            signal_generator.set_power(set_power)
            time.sleep(settling_time)

            measured_power = self._measure_with_retry(power_meter, measurement_times)

            if measured_power is None:
                self.add_power_sweep_point(frequency, set_power, None, None, "测量失败")
                consecutive_failures += 1
                print("警告: 功率计测量失败。")
                if consecutive_failures >= 3:
                    stop_reason = "连续 3 个功率点测量失败，终止扫描"
                    print(f"停止条件: {stop_reason}")
                    break
                continue

            consecutive_failures = 0
            measured_count += 1
            if max_measured is None or measured_power > max_measured:
                max_measured = measured_power
            if min_measured is None or measured_power < min_measured:
                min_measured = measured_power

            compensated_power = measured_power + (attenuator_value if attenuator_enabled else 0.0)
            self.add_power_sweep_point(
                frequency, set_power, measured_power, compensated_power, "正常"
            )
            print(f"测量功率: {measured_power:.2f} dBm, "
                  f"补偿功率: {compensated_power:.2f} dBm")

            if measured_power >= max_measured_power:
                stop_reason = f"测量功率达到保护上限 ({max_measured_power} dBm)"
                print(f"停止条件: {stop_reason}")
                break

        if not keep_output:
            signal_generator.enable_output(False)
            time.sleep(post_close_wait)

        self.add_test_result(
            frequency=frequency,
            total_points=len(power_points),
            measured_points=measured_count,
            max_measured_power=max_measured,
            min_measured_power=min_measured,
            attenuator_enabled=attenuator_enabled,
            attenuation=attenuator_value if attenuator_enabled else 0.0,
            stop_reason=stop_reason,
        )

        print(f"\n{'=' * 60}")
        print(f"扫描完成: {test_config['test_name']}")
        print(f"成功测量点数: {measured_count}/{len(power_points)}")
        if max_measured is not None:
            print(f"最大测量功率: {max_measured:.2f} dBm")
            print(f"最小测量功率: {min_measured:.2f} dBm")
        print(f"停止原因: {stop_reason}")
        print(f"{'=' * 60}")

    def save_results(self, filename):
        """保存测试结果到 Excel。"""
        if not self.test_results:
            print("没有测试结果可保存。")
            return

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "测试总结"
        summary = self.test_results[0]
        rows = [
            ("测试项目", "单频点功率扫描"),
            ("测量频率", f"{summary['frequency']} Hz "
                         f"({format_frequency(summary['frequency'])})"),
            ("扫描点数", summary['total_points']),
            ("成功测量点数", summary['measured_points']),
            ("最大测量功率", summary['max_measured_power']
                        if summary['max_measured_power'] is not None
                        else "无有效数据"),
            ("最小测量功率", summary['min_measured_power']
                        if summary['min_measured_power'] is not None
                        else "无有效数据"),
            ("衰减器补偿", f"{summary['attenuation']} dB"
                        if summary['attenuator_enabled'] else "未启用"),
            ("停止原因", summary['stop_reason']),
            ("测试时间", summary['timestamp']),
        ]
        for row, (label, value) in enumerate(rows, 1):
            ws_summary.cell(row=row, column=1).value = label
            ws_summary.cell(row=row, column=2).value = value
        ws_summary.cell(row=1, column=1).font = Font(bold=True)
        ws_summary.column_dimensions['A'].width = 22
        ws_summary.column_dimensions['B'].width = 40

        ws_detail = wb.create_sheet(title="详细数据")
        headers = [
            '频率 (Hz)', '设定功率 (dBm)', '测量功率 (dBm)',
            '补偿功率 (dBm)', '状态', '时间戳',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row, data in enumerate(self.power_sweep_data, 2):
            ws_detail.cell(row=row, column=1).value = data['frequency']
            ws_detail.cell(row=row, column=2).value = data['set_power']
            ws_detail.cell(row=row, column=3).value = data['measured_power']
            ws_detail.cell(row=row, column=4).value = data['compensated_power']
            ws_detail.cell(row=row, column=5).value = data['status']
            ws_detail.cell(row=row, column=6).value = data['timestamp']

        for col in range(1, len(headers) + 1):
            ws_detail.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = 20

        wb.save(filename)
        print(f"测试结果已保存到: {filename}")
