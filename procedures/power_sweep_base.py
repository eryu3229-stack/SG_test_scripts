import time
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from base_test_procedure import BaseTestProcedure, format_frequency



class PowerSweepBaseProcedure(BaseTestProcedure):
    """最大功率和低频最大功率测试的公共基类"""
    def __init__(self, instrument_manager):
        """初始化测试流程

        Args:
            instrument_manager: 仪器管理器对象
        """
        super().__init__(instrument_manager)
        self.power_sweep_data = []  # 详细的功率扫描数据（用于调试和分析）
        self.csv_sweep_streamer = None

    def add_test_result(self, frequency, max_power, max_measured_power, 
                       attenuation, saturation_point, steps, notes):
        """添加测试结果

        Args:
            frequency: 测量频率
            max_power: 最大可用功率 (dBm)，考虑衰减器补偿后的信号源输出功率
            max_measured_power: 最大测量功率 (dBm)，功率计实际读数
            attenuation: 使用的衰减值 (dB)
            saturation_point: 是否检测到饱和点
            steps: 执行的功率步进数
            notes: 备注信息
        """
        self.test_results.append({
            'frequency': frequency,
            'max_power': max_power,
            'max_measured_power': max_measured_power,
            'attenuation': attenuation,
            'saturation_point': saturation_point,
            'steps': steps,
            'notes': notes,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    def add_power_sweep_point(self, frequency, set_power, measured_power, 
                             actual_power, step_index, status):
        """添加功率扫描数据点

        Args:
            frequency: 频率
            set_power: 设定功率 (dBm)
            measured_power: 测量功率 (dBm)
            actual_power: 实际功率 (dBm) = 测量功率 + 衰减
            step_index: 步进索引
            status: 状态字符串（如'正常', '饱和', '过载', '超限'）
        """
        self.power_sweep_data.append({
            'frequency': frequency,
            'set_power': set_power,
            'measured_power': measured_power,
            'actual_power': actual_power,
            'step_index': step_index,
            'status': status,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        if self.csv_sweep_streamer:
            self.csv_sweep_streamer.append(self.power_sweep_data[-1])

    def start_csv_stream(self, csv_path):
        """开启 CSV 流式写入"""
        from utils.csv_streamer import CsvStreamer
        self.csv_sweep_streamer = CsvStreamer(
            csv_path.replace(".csv", "_sweep.csv"), [
                "frequency", "set_power", "measured_power", "actual_power",
                "step_index", "status", "timestamp",
            ]
        )

    def finish_xlsx(self, xlsx_path):
        """关闭 CSV 流并转为 XLSX"""
        if not self.csv_sweep_streamer:
            return self.save_results(xlsx_path)
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment
        df_sweep = pd.read_csv(self.csv_sweep_streamer.filepath, encoding="utf-8-sig")
        self.csv_sweep_streamer.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "最大功率摘要"
        for col, h in enumerate(["频率 (Hz)", "最大实际功率 (dBm)", "最大测量功率 (dBm)", "衰减值 (dB)", "是否饱和", "步进数", "备注", "时间戳"], 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for r, row in pd.DataFrame(self.test_results).iterrows():
            for c, val in enumerate(row, 1):
                ws.cell(row=r+4, column=c, value=val)
        ws2 = wb.create_sheet(title="详细功率扫描数据")
        for col, h in enumerate(["频率 (Hz)", "设定功率 (dBm)", "测量功率 (dBm)", "实际功率 (dBm)", "步进索引", "状态", "时间戳"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for r, row in df_sweep.iterrows():
            for c, val in enumerate(row, 1):
                ws2.cell(row=r+2, column=c, value=val)
        wb.save(xlsx_path)
        print(f"Excel 已保存: {xlsx_path}")

    def save_results(self, filename):
        """保存测试结果

        Args:
            filename: 保存结果的文件名
        """
        if not self.test_results:
            print("没有测试结果可保存")
            return

        try:
            if filename.endswith('.xlsx'):
                # 保存为Excel格式 - 两个工作表：摘要和详细数据
                wb = openpyxl.Workbook()
                
                # 工作表1: 最大功率摘要
                ws_summary = wb.active
                ws_summary.title = "最大功率摘要"
                
                # 写入表头
                headers = ['频率 (Hz)', '频率显示', '最大实际功率 (dBm)', '最大测量功率 (dBm)', 
                          '衰减值 (dB)', '是否饱和', '功率步进数', '备注', '时间戳']
                for col, header in enumerate(headers, 1):
                    cell = ws_summary.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for row, result in enumerate(self.test_results, 2):
                    ws_summary.cell(row=row, column=1).value = result['frequency']
                    ws_summary.cell(row=row, column=2).value = format_frequency(result['frequency'])
                    ws_summary.cell(row=row, column=3).value = result['max_power']
                    ws_summary.cell(row=row, column=4).value = result['max_measured_power']
                    ws_summary.cell(row=row, column=5).value = result['attenuation']
                    ws_summary.cell(row=row, column=6).value = '是' if result['saturation_point'] else '否'
                    ws_summary.cell(row=row, column=7).value = result['steps']
                    ws_summary.cell(row=row, column=8).value = result['notes']
                    ws_summary.cell(row=row, column=9).value = result['timestamp']
                
                # 调整列宽
                for col in range(1, len(headers) + 1):
                    ws_summary.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # 工作表2: 详细功率扫描数据
                ws_detail = wb.create_sheet(title="详细功率扫描数据")
                
                # 写入表头
                detail_headers = ['频率 (Hz)', '设定功率 (dBm)', '测量功率 (dBm)', 
                                '实际功率 (dBm)', '步进索引', '状态', '时间戳']
                for col, header in enumerate(detail_headers, 1):
                    cell = ws_detail.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for row, data in enumerate(self.power_sweep_data, 2):
                    ws_detail.cell(row=row, column=1).value = data['frequency']
                    ws_detail.cell(row=row, column=2).value = data['set_power']
                    ws_detail.cell(row=row, column=3).value = data['measured_power']
                    ws_detail.cell(row=row, column=4).value = data['actual_power']
                    ws_detail.cell(row=row, column=5).value = data['step_index']
                    ws_detail.cell(row=row, column=6).value = data['status']
                    ws_detail.cell(row=row, column=7).value = data['timestamp']
                
                # 调整列宽
                for col in range(1, len(detail_headers) + 1):
                    ws_detail.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                wb.save(filename)
                print(f"测试结果已保存到Excel文件: {filename}")
                
            else:
                # 保存为CSV格式 - 两个文件
                base_name = filename.replace('.csv', '')
                summary_file = f"{base_name}_summary.csv"
                detail_file = f"{base_name}_detail.csv"
                
                # 保存摘要
                with open(summary_file, 'w', newline='') as csvfile:
                    fieldnames = ['frequency', 'max_power', 'max_measured_power', 
                                 'attenuation', 'saturation_point', 'steps', 'notes', 'timestamp']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for result in self.test_results:
                        writer.writerow(result)
                
                # 保存详细数据
                with open(detail_file, 'w', newline='') as csvfile:
                    fieldnames = ['frequency', 'set_power', 'measured_power', 
                                 'actual_power', 'step_index', 'status', 'timestamp']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for data in self.power_sweep_data:
                        writer.writerow(data)
                
                print(f"测试结果已保存到CSV文件:")
                print(f"  摘要: {summary_file}")
                print(f"  详细数据: {detail_file}")
                
        except Exception as e:
            print(f"保存测试结果失败: {e}")