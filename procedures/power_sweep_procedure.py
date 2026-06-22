import time
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from base_test_procedure import BaseTestProcedure, format_frequency


class TestProcedure(BaseTestProcedure):
    """功率扫描测试流程类"""

    def __init__(self, instrument_manager):
        """初始化测试流程"""
        super().__init__(instrument_manager)
        self.test_prepared = False

    def add_test_result(self, frequency, set_power, measured_power, attenuator_value=0):
        """添加测试结果"""
        compensated_power = measured_power + attenuator_value if attenuator_value > 0 else measured_power
        self.test_results.append({
            'frequency': frequency,
            'set_power': set_power,
            'measured_power': measured_power,
            'compensated_power': compensated_power,
            'attenuator_value': attenuator_value,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        if self.csv_streamer:
            self.csv_streamer.append(self.test_results[-1])

    def prepare_test(self, signal_generator, power_meter):
        """测试前准备步骤"""
        print("等待功率计归零稳定...")
        time.sleep(2.0)
        print("开启测试准备...")
        signal_generator.enable_output(False)
        time.sleep(0.5)
        self.test_prepared = True
        print("测试准备完成")

    def run_test(self, signal_generator, power_meter, test_config, keep_output=False):
        """运行测试"""
        print(f"开始测试: {test_config['test_name']}")

        signal_generator.set_frequency(test_config['frequency'])
        signal_generator.set_power(test_config['power'])
        signal_generator.enable_output(True)

        settling_time = test_config.get('settling_time', 1.5)
        time.sleep(settling_time)

        power_meter.set_frequency(test_config['frequency'])
        pm_settling_time = test_config.get('pm_settling_time', 0.5)
        time.sleep(pm_settling_time)

        measurement_times = test_config.get('measurement_times', 5)
        measured_power = power_meter.measure_power(times=measurement_times)

        attenuator_value = 0
        if test_config.get('attenuator_enabled', False):
            attenuator_value = test_config.get('attenuator_value', 0)

        self.add_test_result(test_config['frequency'], test_config['power'], measured_power, attenuator_value)

        if not keep_output:
            signal_generator.enable_output(False)
            post_close_wait = test_config.get('post_close_wait', 0.1)
            time.sleep(post_close_wait)

        print(f"测试完成: {test_config['test_name']}")

    def start_csv_stream(self, csv_path):
        """开启 CSV 流式写入"""
        from utils.csv_streamer import CsvStreamer
        self.csv_streamer = CsvStreamer(csv_path, [
            "frequency", "set_power", "measured_power", "compensated_power",
            "attenuator_value", "timestamp",
        ])

    def finish_xlsx(self, xlsx_path, test_configs=None):
        """关闭 CSV 流并转为 XLSX"""
        if not self.csv_streamer:
            return self.save_results(xlsx_path)
        import pandas as pd
        df = pd.read_csv(self.csv_streamer.filepath, encoding="utf-8-sig")
        self.csv_streamer.close()
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试数据"
        ws.cell(row=1, column=1, value="测试总结").font = Font(bold=True, size=14)
        for col, h in enumerate(["测量频率", "设定功率", "实际功率", "补偿功率", "衰减器值", "时间戳"], 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for r, row in df.iterrows():
            for c, val in enumerate(row, 1):
                ws.cell(row=r+4, column=c, value=val)
        wb.save(xlsx_path)
        print(f"Excel 已保存: {xlsx_path}")

    def save_results(self, filename, test_configs=None):
        """保存测试结果"""
        if not self.test_results:
            print("没有测试结果可保存")
            return

        try:
            if not filename.endswith('.xlsx'):
                filename = filename.rsplit('.', 1)[0] + '.xlsx'
                print(f"自动转换为Excel格式: {filename}")
            
            wb = openpyxl.Workbook()
            summary_ws = wb.active
            summary_ws.title = "测试总结"
            
            summary_ws.cell(row=1, column=1).value = "测试总结"
            summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            
            summary_ws.cell(row=3, column=1).value = "测试项目"
            summary_ws.cell(row=3, column=2).value = "功率扫描测试"
            summary_ws.cell(row=4, column=1).value = "测试时间"
            summary_ws.cell(row=4, column=2).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            summary_ws.cell(row=5, column=1).value = "测试点数"
            summary_ws.cell(row=5, column=2).value = len(self.test_results)
            
            summary_ws.cell(row=7, column=1).value = "测试配置"
            summary_ws.cell(row=7, column=1).font = Font(bold=True)
            
            if test_configs and len(test_configs) > 0:
                config = test_configs[0]
                row = 8
                for key, label, unit in [
                    ('settling_time', '信号源稳定时间', '秒'),
                    ('pm_settling_time', '功率计频率切换稳定时间', '秒'),
                    ('post_close_wait', '信号源关闭后等待时间', '秒'),
                ]:
                    summary_ws.cell(row, column=1).value = label
                    summary_ws.cell(row, column=2).value = config.get(key, 'N/A')
                    summary_ws.cell(row, column=3).value = unit
                    row += 1
                
                summary_ws.cell(row, column=1).value = '功率计测量次数'
                summary_ws.cell(row, column=2).value = config.get('measurement_times', 5)
                row += 1
                
                summary_ws.cell(row, column=1).value = '衰减器启用'
                summary_ws.cell(row, column=2).value = '是' if config.get('attenuator_enabled', False) else '否'
                row += 1
                
                if config.get('attenuator_enabled', False):
                    summary_ws.cell(row, column=1).value = '衰减器类型'
                    if config.get('attenuator_freq_dependent', False):
                        summary_ws.cell(row, column=2).value = '频率相关'
                    else:
                        summary_ws.cell(row, column=2).value = '固定值'
                        row += 1
                        summary_ws.cell(row, column=1).value = '衰减值'
                        summary_ws.cell(row, column=2).value = config.get('attenuator_value', 0)
                        summary_ws.cell(row, column=3).value = 'dB'
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 10
            
            data_ws = wb.create_sheet(title="测试数据")
            headers = ['测量频率', '设定功率', '实际功率', '补偿功率', '衰减器值', '时间戳']
            for col, header in enumerate(headers, 1):
                cell = data_ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            for row, result in enumerate(self.test_results, 2):
                data_ws.cell(row=row, column=1).value = result['frequency']
                data_ws.cell(row=row, column=2).value = result['set_power']
                data_ws.cell(row=row, column=3).value = result['measured_power']
                data_ws.cell(row=row, column=4).value = result['compensated_power']
                data_ws.cell(row=row, column=5).value = result['attenuator_value']
                data_ws.cell(row=row, column=6).value = result['timestamp']
            
            for col in range(1, len(headers) + 1):
                data_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
            
            wb.save(filename)
            print(f"测试结果已保存到: {filename}")
        except Exception as e:
            print(f"保存测试结果失败: {e}")